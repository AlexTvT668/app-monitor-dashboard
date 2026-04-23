#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小米渠道 2025 vs 2026 1-4 月同比分析
输入：两份运营数据列表导出（xlsx）
输出：analytics/xiaomi_insights.json  供前端业务数据分析模块使用
"""
from __future__ import annotations
import json, os, sys
from collections import defaultdict
from datetime import date, timedelta
from openpyxl import load_workbook

FILE_25 = '/Users/alex/Downloads/数据中心-运营数据列表导出-20260423 (1).xlsx'  # 25 年数据
FILE_26 = '/Users/alex/Downloads/数据中心-运营数据列表导出-20260423.xlsx'       # 26 年数据
OUT_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'xiaomi_insights.json')

# —— 口径 ——
# 日期格式 YYYYMMDD(int)；列顺序见表头
COL = {'date':0,'supplier':1,'channel':2,'appid':3,'app':4,
       'new':5,'dau':6,'pay_user':7,'revenue':8,
       'arppu':9,'arpu':10,'pay_rate':11,'r1':12,'r3':13,'r7':14}

def to_num(v):
    """把 '31390.90' / '2.39%' / '-' / '0.43' 这种混合统一转 float；无效返回 0.0"""
    if v is None: return 0.0
    if isinstance(v, (int,float)): return float(v)
    s = str(v).strip()
    if s in ('','-','—','N/A','nan','None'): return 0.0
    if s.endswith('%'):
        try: return float(s[:-1])/100.0
        except: return 0.0
    try: return float(s)
    except: return 0.0

def d_int(i:int)->date:
    return date(i//10000, (i//100)%100, i%100)

def load_rows(path:str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for idx, r in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0: continue
        if r[0] is None: continue
        rows.append(r)
    wb.close()
    return rows

def aggregate(rows, start:int, end:int):
    """返回 (totals, per_day, per_app)；start/end 为 YYYYMMDD 闭区间"""
    totals = {'revenue':0.0,'new':0,'dau_sum':0,'pay_user_sum':0,'days':0}
    per_day = defaultdict(lambda: {'revenue':0.0,'new':0,'dau':0,'pay':0})
    per_app = defaultdict(lambda: {'revenue':0.0,'new':0,'dau':0,'pay':0})
    seen_dates = set()
    for r in rows:
        d = r[COL['date']]
        if d < start or d > end: continue
        seen_dates.add(d)
        rev = to_num(r[COL['revenue']])
        nu  = int(to_num(r[COL['new']]))
        du  = int(to_num(r[COL['dau']]))
        pu  = int(to_num(r[COL['pay_user']]))
        app = r[COL['app']] or '(未知)'
        totals['revenue'] += rev
        totals['new']     += nu
        totals['dau_sum'] += du
        totals['pay_user_sum'] += pu
        per_day[d]['revenue'] += rev
        per_day[d]['new']     += nu
        per_day[d]['dau']     += du
        per_day[d]['pay']     += pu
        pa = per_app[app]
        pa['revenue'] += rev
        pa['new']     += nu
        pa['dau']     += du
        pa['pay']     += pu
    totals['days'] = len(seen_dates)
    return totals, per_day, per_app

def yoy(a, b):
    """b 为基期。返回 (a-b)/b，基期为 0 时返回 None"""
    if b in (0,0.0,None): return None
    return (a-b)/b

def top_n(per_app, n=10, key='revenue'):
    return sorted(per_app.items(), key=lambda kv: kv[1][key], reverse=True)[:n]

def fmt_day(i:int)->str:
    d = d_int(i)
    return f'{d.month:02d}-{d.day:02d}'

def ymd_range(d0:date, d1:date):
    """生成 YYYYMMDD int 列表"""
    out = []
    cur = d0
    while cur <= d1:
        out.append(cur.year*10000 + cur.month*100 + cur.day)
        cur += timedelta(days=1)
    return out

def main():
    print('[load] 25 data ...')
    rows25 = load_rows(FILE_25)
    print(f'  rows={len(rows25)}')
    print('[load] 26 data ...')
    rows26 = load_rows(FILE_26)
    print(f'  rows={len(rows26)}')

    # ——— 1. 对齐全量期间：01-01 ~ 04-22 ———
    PERIOD_25 = (20250101, 20250422)
    PERIOD_26 = (20260101, 20260422)
    t25, d25, a25 = aggregate(rows25, *PERIOD_25)
    t26, d26, a26 = aggregate(rows26, *PERIOD_26)

    # ——— 2. Q1 期间：01-01 ~ 03-31 ———
    Q1_25 = (20250101, 20250331)
    Q1_26 = (20260101, 20260331)
    q25, qd25, qa25 = aggregate(rows25, *Q1_25)
    q26, qd26, qa26 = aggregate(rows26, *Q1_26)

    # ——— 3. 春节窗口 ———
    # 2025 春节 1/29；2026 春节 2/17；窗口=春节前 3 天 ~ 春节后 7 天（共 11 天）
    def window(cny:date):
        start = cny - timedelta(days=3)
        end   = cny + timedelta(days=7)
        return start.year*10000+start.month*100+start.day, end.year*10000+end.month*100+end.day
    CNY25_S, CNY25_E = window(date(2025,1,29))
    CNY26_S, CNY26_E = window(date(2026,2,17))
    c25, cd25, ca25 = aggregate(rows25, CNY25_S, CNY25_E)
    c26, cd26, ca26 = aggregate(rows26, CNY26_S, CNY26_E)

    # ——— 4. Top10 流水游戏（基于对齐全量期间） ———
    top25 = top_n(a25, 10)
    top26 = top_n(a26, 10)
    # 26 年 Top10 阵容 + 25 年同期对比
    top26_with_25 = []
    for app, v in top26:
        v25 = a25.get(app, {'revenue':0.0,'new':0,'dau':0,'pay':0})
        top26_with_25.append({
            'app': app,
            'rev26': round(v['revenue'],2),
            'rev25': round(v25['revenue'],2),
            'yoy': yoy(v['revenue'], v25['revenue']),
            'new26': v['new'], 'new25': v25['new'],
            'dau26_avg': round(v['dau']/max(t26['days'],1),0),
            'dau25_avg': round(v25['dau']/max(t25['days'],1),0),
            'rank25': next((i+1 for i,(a,_) in enumerate(top_n(a25, 999)) if a==app), None),
        })

    # 新入榜 / 掉榜
    apps25 = [a for a,_ in top25]
    apps26 = [a for a,_ in top26]
    new_in   = [a for a in apps26 if a not in apps25]
    dropped  = [a for a in apps25 if a not in apps26]

    # ——— 5. 每日流水时序（用于走势图） ———
    def daily_series(per_day, year:int, s:int, e:int):
        out = []
        for ymd in ymd_range(d_int(s), d_int(e)):
            v = per_day.get(ymd, {})
            out.append({
                'md': fmt_day(ymd),
                'revenue': round(v.get('revenue',0.0),2),
                'new': v.get('new',0),
                'dau': v.get('dau',0),
                'pay': v.get('pay',0),
            })
        return out
    series25 = daily_series(d25, 2025, *PERIOD_25)
    series26 = daily_series(d26, 2026, *PERIOD_26)

    # ——— 6. 汇总衍生指标 ———
    def derive(t):
        days = max(t['days'],1)
        return {
            'revenue': round(t['revenue'],2),
            'revenue_wan': round(t['revenue']/10000, 2),  # 万元
            'new': t['new'],
            'dau_avg': round(t['dau_sum']/days, 0),
            'pay_avg': round(t['pay_user_sum']/days, 0),
            'arppu':   round(t['revenue']/t['pay_user_sum'], 2) if t['pay_user_sum'] else 0,
            'pay_rate': round(t['pay_user_sum']/t['dau_sum'], 4) if t['dau_sum'] else 0,
            'days': t['days'],
        }

    overall = {
        'period_label': '2025/2026 · 01-01 ~ 04-22 同窗口对齐',
        'y25': derive(t25), 'y26': derive(t26),
        'yoy': {
            'revenue': yoy(t26['revenue'], t25['revenue']),
            'new':     yoy(t26['new'],     t25['new']),
            'dau':     yoy(t26['dau_sum'], t25['dau_sum']),
            'pay':     yoy(t26['pay_user_sum'], t25['pay_user_sum']),
        }
    }
    q1 = {
        'period_label': 'Q1（01-01 ~ 03-31）同比',
        'y25': derive(q25), 'y26': derive(q26),
        'yoy': {
            'revenue': yoy(q26['revenue'], q25['revenue']),
            'new':     yoy(q26['new'],     q25['new']),
            'dau':     yoy(q26['dau_sum'], q25['dau_sum']),
            'pay':     yoy(q26['pay_user_sum'], q25['pay_user_sum']),
        }
    }
    cny = {
        'period_label_25': f'25 春节窗口 01-26 ~ 02-05（春节 1/29）',
        'period_label_26': f'26 春节窗口 02-14 ~ 02-24（春节 2/17）',
        'y25': derive(c25), 'y26': derive(c26),
        'yoy': {
            'revenue': yoy(c26['revenue'], c25['revenue']),
            'new':     yoy(c26['new'],     c25['new']),
            'dau':     yoy(c26['dau_sum'], c25['dau_sum']),
            'pay':     yoy(c26['pay_user_sum'], c25['pay_user_sum']),
        },
        # 春节窗口 Top5
        'top5_25': [{'app':a,'rev':round(v['revenue'],2)} for a,v in top_n(ca25,5)],
        'top5_26': [{'app':a,'rev':round(v['revenue'],2)} for a,v in top_n(ca26,5)],
    }

    out = {
        'vendor': '小米',
        'supplier': '北京瓦力网络科技有限公司',
        'channel_count': 5,
        'channel_codes': [10033159, 10159208, 10027723, 10027724, 10003898],
        'generated_at': '2026-04-23',
        'overall': overall,
        'q1': q1,
        'cny': cny,
        'top10_26': top26_with_25,
        'top10_25': [{'app':a,'rev':round(v['revenue'],2),'new':v['new']} for a,v in top25],
        'movers': {'new_in': new_in, 'dropped': dropped},
        'series25': series25,
        'series26': series26,
    }

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f'[done] wrote {OUT_JSON}')

    # 打印核心结论供人工核查
    print('\n========== 核心结论 ==========')
    print('—— 全窗口（01-01 ~ 04-22） ——')
    print(f"  25 年流水: {overall['y25']['revenue_wan']:>10,.2f} 万")
    print(f"  26 年流水: {overall['y26']['revenue_wan']:>10,.2f} 万")
    print(f"  同比:     {overall['yoy']['revenue']*100:+.2f}%" if overall['yoy']['revenue'] is not None else '  同比: N/A')
    print(f"  25 新增:  {overall['y25']['new']:>12,}   26 新增: {overall['y26']['new']:>12,}   同比: {overall['yoy']['new']*100:+.2f}%")
    print(f"  25 日均DAU: {overall['y25']['dau_avg']:>10,.0f}  26 日均DAU: {overall['y26']['dau_avg']:>10,.0f}  同比: {overall['yoy']['dau']*100:+.2f}%")
    print(f"  25 ARPPU: {overall['y25']['arppu']:.2f}  26 ARPPU: {overall['y26']['arppu']:.2f}")
    print(f"  25 付费率: {overall['y25']['pay_rate']*100:.2f}%  26 付费率: {overall['y26']['pay_rate']*100:.2f}%")
    print('—— Q1 同比 ——')
    print(f"  25 Q1 流水: {q1['y25']['revenue_wan']:>10,.2f} 万   26 Q1 流水: {q1['y26']['revenue_wan']:>10,.2f} 万   同比: {q1['yoy']['revenue']*100:+.2f}%")
    print('—— 春节窗口 ——')
    print(f"  25 春节: {cny['y25']['revenue_wan']:,.2f} 万   26 春节: {cny['y26']['revenue_wan']:,.2f} 万   同比: {cny['yoy']['revenue']*100:+.2f}%")
    print('—— Top10（26 年） ——')
    for i, it in enumerate(top26_with_25, 1):
        y = f"{it['yoy']*100:+.1f}%" if it['yoy'] is not None else 'N/A'
        print(f"  #{i:2d}  {it['app']:<20}  26:{it['rev26']/10000:>9,.1f}万  25:{it['rev25']/10000:>9,.1f}万  同比:{y:>8}  25排名:{it['rank25']}")
    print('—— 新入榜/掉榜 ——')
    print(f"  new_in:  {new_in}")
    print(f"  dropped: {dropped}")

if __name__ == '__main__':
    main()
