"""
解析渠道月度规划 Excel（示例：小米渠道规划.xlsx）
结构：类别 / 事项 / 进展 / 规划
输出：analytics/<vendor>_plan.json
"""
import json
from pathlib import Path
import openpyxl


def parse_sheet(ws):
    """解析单张规划 sheet，返回 [{category,item,progress,plan}]。
    类别列为空时沿用上一个类别（合并单元格场景）。"""
    rows = []
    cur_cat = ""
    for r in range(2, ws.max_row + 1):
        def cell(c):
            v = ws.cell(r, c).value
            return str(v).strip() if v is not None else ""
        cat, item, prog, plan = cell(1), cell(2), cell(3), cell(4)
        if not any([cat, item, prog, plan]):
            continue
        if cat:
            cur_cat = cat
        rows.append({
            "category": cur_cat,
            "item": item,
            "progress": prog,
            "plan": plan,
        })
    return rows


def parse_workbook(xlsx_path: Path, vendor_key: str, vendor_name: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    months = []
    for sn in wb.sheetnames:
        if "月规划" not in sn:
            continue
        ws = wb[sn]
        if ws.max_row < 2 or ws.max_column < 4:
            continue
        rows = parse_sheet(ws)
        if not rows:
            continue
        m_label = sn.replace("规划", "").strip()
        months.append({
            "sheet_name": sn,
            "month": m_label,
            "label": m_label + "规划",
            "rows": rows,
            "row_count": len(rows),
        })
    # 类别汇总
    cat_stats = {}
    for m in months:
        for r in m["rows"]:
            cat_stats.setdefault(r["category"], 0)
            cat_stats[r["category"]] += 1
    return {
        "vendor_key": vendor_key,
        "vendor_name": vendor_name,
        "source_file": xlsx_path.name,
        "months": months,
        "category_counts": cat_stats,
    }


def main():
    xlsx = Path("/Users/alex/Downloads/小米渠道规划.xlsx")
    out = Path(__file__).resolve().parent / "xiaomi_plan.json"
    data = parse_workbook(xlsx, vendor_key="xiaomi", vendor_name="小米")
    out.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    for m in data["months"]:
        print(f"{m['label']}: {m['row_count']} rows")
    print(f"categories: {data['category_counts']}")
    print(f"-> {out}")


if __name__ == "__main__":
    main()
